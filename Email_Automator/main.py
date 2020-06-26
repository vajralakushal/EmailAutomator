# smpt-mail.outlook.com
#script process: 
#read inbox -> parse each email -> put email information into excel sheet -> read excel sheet 
# -> extrapolate information -> send email to customer 
import imaplib
import base64
import os
import email
import xlrd
import openpyxl
from email.message import EmailMessage

def parse_email(email):
    email_body = email.get_content()
    email_body = email_body.replace("\n\n", " ")
    email_body = email_body.replace("\n", " ")
    email_body = email_body.replace(": ", ":")
    param_dict = {}
    key = ""
    value = ""
    onValue = False
    onKey = True
    onComments = False
    for character in email_body:
        if onKey:
            if character == ':':
                if key == "Comments":
                    onComments = True
                onValue = True
                onKey = False
                continue
            else:
                key = key + character
        elif onValue:
            if not onComments:
                if character == ' ':
                    onValue = False
                    onKey = True
                    param_dict[key] = value
                    key = ""
                    value = ""
                else:
                    value = value + character
            elif email_body.endswith(character):
                value = value + " "
            else:
                value = value + character
    param_dict[key] = value            
    return param_dict

imap_clients = {
    "outlook" : ("imap-mail.outlook.com", 993)
}

loginFile = open("credentials.txt", "r")
connectionInformation = {}
for line in loginFile:
    key = line.split(":")[0]
    value = line.split(":")[1].rstrip("\n")
    connectionInformation[key] = value

client = connectionInformation["client"]

#Connect and login to outlook mailbox
mailbox = imaplib.IMAP4_SSL(imap_clients[client][0], imap_clients[client][1])
print("successful connection to outlook\n")
username = connectionInformation["username"]
password = connectionInformation["password"]
mailbox.login(username, password)
print("successful login\n")

#Set the current mailbox to search from and start going thru unread emails
mailbox.select('Inbox')
#Container to hold unread emails
unread_emails = []
type, data = mailbox.search(None, '(UNSEEN)')
mail_ids = data[0]
id_list = mail_ids.split()
temp = data[0].split()
if type == "OK":
    for num in temp:
        typ, data = mailbox.fetch(num, '(RFC822)' )
        raw_email = data[0][1]
        for response_part in data:
            if isinstance(response_part, tuple):
                #Create new Email Message object and add it to unread_emails array
                msg = email.message_from_string(response_part[1].decode('utf-8'))

                unread_email = EmailMessage()
                unread_email['Subject'] = msg['subject']
                unread_email['From'] = msg['from']
                unread_email.set_content(msg.get_payload()[0].get_payload())

                unread_emails.append(unread_email)
print(f"Unread emails fetched:{len(unread_emails)}")

#Close connection to outlook and logout
mailbox.close()
mailbox.logout()
print("Closing connection and logging off")

workbook = openpyxl.load_workbook("templates.xlsx")
customers = workbook["Customers"]

#Parse and create param_dict + key_dict
#Insert information into excel file
for email in unread_emails:
    param_dict = parse_email(email)

    key_dict = []
    for row in customers.iter_cols(min_row=1, max_col=customers.max_column, max_row=2, values_only=True):
        key_dict.append(row[0])

    new_customer_row = customers.max_row + 1
    for current_col in range(customers.max_column):
        if key_dict[current_col] in param_dict:
            customers.cell(row=new_customer_row, column=current_col + 1).value = param_dict[key_dict[current_col]]
